# app.py
# Integrated Financial Performance & Governance Dashboard
# Combines:
# 1) Business_Financial_KPI_Dashboard_Template.xlsx (KPI Template / ledger driven)
# 2) SDE & SOE Graph (2).xlsx (SDE/SOE/CDOE year-wise trends + governance)
#
# Run:
#   pip install streamlit pandas openpyxl plotly
#   streamlit run app.py

import os
import re
from datetime import datetime, date
from calendar import monthrange

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

# -----------------------------
# CONFIG
# -----------------------------
st.set_page_config(
    page_title="Integrated Financial Performance & Governance Dashboard",
    layout="wide",
)

DEFAULT_KPI_PATH = r"/mnt/data/Business_Financial_KPI_Dashboard_Template.xlsx"
DEFAULT_SDE_PATH = r"/mnt/data/SDE & SOE Graph (2).xlsx"

MODEL_MAP_IMG_PATHS = [
    r"/mnt/data/6235b073-759e-4de2-bf1f-dea0a0250523.png",  # overview/model map screenshot
    r"/mnt/data/59153642-a6f9-40fd-b370-41ae5a98bf11.png",
    r"/mnt/data/a8d0c9a7-19eb-40db-b2cc-1d5f628f6f95.png",
]

INR_SYMBOL = "₹"


# -----------------------------
# HELPERS
# -----------------------------
def inr(x):
    """Format INR with commas; no truncation."""
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return "—"
    try:
        x = float(x)
    except Exception:
        return str(x)
    sign = "-" if x < 0 else ""
    x = abs(x)
    return f"{sign}{INR_SYMBOL}{x:,.0f}"


def pct(x, digits=1):
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return "—"
    try:
        return f"{100*float(x):.{digits}f}%"
    except Exception:
        return "—"


def safe_div(a, b):
    try:
        if b in (0, 0.0) or (isinstance(b, float) and np.isnan(b)):
            return np.nan
        return a / b
    except Exception:
        return np.nan


def month_start_end(month_dt: pd.Timestamp):
    ms = pd.Timestamp(month_dt.year, month_dt.month, 1)
    me = pd.Timestamp(month_dt.year, month_dt.month, monthrange(month_dt.year, month_dt.month)[1])
    return ms, me


def to_month_ts(x):
    """Force month-level Timestamp."""
    if isinstance(x, pd.Timestamp):
        return pd.Timestamp(x.year, x.month, 1)
    try:
        t = pd.to_datetime(x)
        return pd.Timestamp(t.year, t.month, 1)
    except Exception:
        return pd.NaT


def rag_from_threshold(value, green_max=None, green_min=None, amber_band=0.1, higher_is_bad=True):
    """
    Generic RAG:
      - higher_is_bad=True + green_max:
          GREEN if value <= (1-amber_band)*green_max
          AMBER if within ((1-amber_band)*green_max, green_max]
          RED   if > green_max
      - higher_is_bad=False + green_min:
          GREEN if value >= (1+amber_band)*green_min
          AMBER if within [green_min, (1+amber_band)*green_min)
          RED   if < green_min
    """
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return "AMBER", "Insufficient data (missing/zero denominator)."

    if higher_is_bad:
        if green_max is None:
            return "AMBER", "No threshold set."
        if value > green_max:
            return "RED", "Breached limit."
        if value > (1 - amber_band) * green_max:
            return "AMBER", "Near limit."
        return "GREEN", "Within comfort zone."
    else:
        if green_min is None:
            return "AMBER", "No threshold set."
        if value < green_min:
            return "RED", "Below minimum."
        if value < (1 + amber_band) * green_min:
            return "AMBER", "Barely above minimum."
        return "GREEN", "Comfortably above minimum."


def rag_box(color, text):
    color_map = {
        "RED":   ("#b42318", "#ffffff"),
        "AMBER": ("#b54708", "#ffffff"),
        "GREEN": ("#067647", "#ffffff"),
        "BLUE":  ("#175cd3", "#ffffff"),
        "GRAY":  ("#475467", "#ffffff"),
    }
    bg, fg = color_map.get(color, ("#475467", "#ffffff"))
    st.markdown(
        f"""
        <div style="background:{bg};color:{fg};padding:14px 16px;border-radius:10px;
                    font-weight:700;margin:8px 0;line-height:1.25">
            {text}
        </div>
        """,
        unsafe_allow_html=True,
    )


def card(title, value, subtitle=None):
    st.markdown(
        f"""
        <div style="border:1px solid #EAECF0;border-radius:14px;padding:14px 16px;background:#ffffff">
            <div style="font-size:12px;color:#667085;font-weight:700">{title}</div>
            <div style="font-size:30px;color:#101828;font-weight:900;margin-top:6px">{value}</div>
            <div style="font-size:12px;color:#667085;margin-top:6px">{subtitle or ""}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def plot_bar_with_labels(df, x, y, color=None, title=None, yaxis_title=None):
    fig = px.bar(df, x=x, y=y, color=color, text=y, title=title)
    fig.update_traces(texttemplate="%{text:,.0f}", textposition="outside", cliponaxis=False)
    fig.update_layout(
        height=520,
        margin=dict(l=20, r=20, t=60, b=40),
        yaxis_title=yaxis_title,
        xaxis_title="",
        legend_title_text="",
    )
    return fig


def plot_grouped_bar_with_labels(df, x, y, group, title=None, yaxis_title=None):
    fig = px.bar(df, x=x, y=y, color=group, barmode="group", text=y, title=title)
    fig.update_traces(texttemplate="%{text:,.0f}", textposition="outside", cliponaxis=False)
    fig.update_layout(
        height=560,
        margin=dict(l=20, r=20, t=60, b=40),
        yaxis_title=yaxis_title,
        xaxis_title="",
        legend_title_text="",
    )
    return fig


def narrative_block(title, bullets):
    st.markdown(f"### {title}")
    for b in bullets:
        st.markdown(f"- {b}")


# -----------------------------
# FILE LOADERS
# -----------------------------
@st.cache_data(show_spinner=False)
def load_kpi_template(excel_bytes_or_path):
    # Transactions is the only sheet we truly need for correct computation
    tx = pd.read_excel(excel_bytes_or_path, sheet_name="Transactions")
    # Normalize columns
    tx.columns = [str(c).strip() for c in tx.columns]
    required = ["Date", "Entity", "AccountType", "Category", "Counterparty", "Amount", "CashFlag", "IntercompanyFlag"]
    missing = [c for c in required if c not in tx.columns]
    if missing:
        raise ValueError(f"KPI template Transactions sheet missing columns: {missing}")

    tx["Date"] = pd.to_datetime(tx["Date"], errors="coerce")
    tx["Entity"] = tx["Entity"].astype(str).str.strip()
    tx["AccountType"] = tx["AccountType"].astype(str).str.strip()
    tx["Category"] = tx["Category"].astype(str).str.strip()
    tx["Counterparty"] = tx["Counterparty"].astype(str).str.strip()
    tx["CashFlag"] = tx["CashFlag"].astype(str).str.strip()
    tx["IntercompanyFlag"] = tx["IntercompanyFlag"].astype(str).str.strip()

    tx["Amount"] = pd.to_numeric(tx["Amount"], errors="coerce").fillna(0.0)

    # Enforce sign convention:
    # Revenue (+), all other listed types (-)
    tx["Amount_norm"] = tx["Amount"].copy()
    is_rev = tx["AccountType"].str.upper().eq("REVENUE")
    tx.loc[is_rev, "Amount_norm"] = tx.loc[is_rev, "Amount"].abs()
    tx.loc[~is_rev, "Amount_norm"] = -tx.loc[~is_rev, "Amount"].abs()

    tx["Month"] = tx["Date"].apply(to_month_ts)
    return tx


@st.cache_data(show_spinner=False)
def load_sde_soe_tables(excel_bytes_or_path):
    # We extract clean tables from specific sheets with merged headers
    def extract_table(sheet_name, header_row_idx):
        raw = pd.read_excel(excel_bytes_or_path, sheet_name=sheet_name, header=None)
        header = raw.iloc[header_row_idx].astype(str).tolist()
        df = raw.iloc[header_row_idx + 1 :].copy()
        df.columns = header
        df = df.dropna(how="all")
        # Drop fully empty columns
        df = df.loc[:, ~df.columns.str.contains(r"^Unnamed", na=False)]
        # Strip column names
        df.columns = [str(c).strip().replace("\n", " ") for c in df.columns]
        return df

    # CDOE(26.02.26): header row at 4 (0-index), from your file inspection it starts at row where "Sr. No."
    # We find it dynamically for safety.
    def find_header_row(sheet_name, needle="Sr. No."):
        raw = pd.read_excel(excel_bytes_or_path, sheet_name=sheet_name, header=None)
        for i in range(len(raw)):
            row = raw.iloc[i].astype(str).tolist()
            if any(needle.lower() in str(x).lower() for x in row):
                return i
        return None

    # Extract CDOE table
    cd_sheet = "CDOE(26.02.26)"
    cd_hr = find_header_row(cd_sheet, "Sr. No.")
    cd = extract_table(cd_sheet, cd_hr)

    # Extract combined table
    comb_sheet = "26.02.26"
    comb_hr = find_header_row(comb_sheet, "Sr. No.")
    comb = extract_table(comb_sheet, comb_hr)

    # Clean / standardize
    def clean_year(df):
        if "Year" not in df.columns:
            # try find any column containing 'Year'
            ycol = next((c for c in df.columns if "year" in c.lower()), None)
            if ycol:
                df = df.rename(columns={ycol: "Year"})
        df["Year"] = df["Year"].astype(str).str.strip()
        return df

    cd = clean_year(cd)
    comb = clean_year(comb)

    # Numeric conversions for likely columns
    for df in [cd, comb]:
        for c in df.columns:
            if c.lower() in {"sr. no.", "sr no", "sr.no."}:
                continue
            if c.lower() == "year":
                continue
            df[c] = pd.to_numeric(df[c], errors="coerce")

    return {"CDOE": cd, "COMBINED": comb}


# -----------------------------
# KPI COMPUTATION
# -----------------------------
def compute_kpi_for_entity_month(tx, entity, month_ts):
    ms, me = month_start_end(month_ts)
    days = (me - ms).days + 1

    df = tx[(tx["Entity"] == entity) & (tx["Date"] >= ms) & (tx["Date"] <= me)].copy()

    # P&L components
    def sum_type(t):
        return df.loc[df["AccountType"].str.upper().eq(t.upper()), "Amount_norm"].sum()

    revenue = sum_type("Revenue")
    cogs = sum_type("COGS")
    opex = sum_type("Opex")
    interest = sum_type("Interest")
    tax = sum_type("Tax")

    gross_profit = revenue + cogs
    ebitda = gross_profit + opex
    net_profit = ebitda + interest + tax

    # Cash flow (cash only)
    ocf = df.loc[df["CashFlag"].str.upper().eq("CASH"), "Amount_norm"].sum()

    # Intercompany
    ic_sum = df.loc[df["IntercompanyFlag"].str.upper().eq("YES"), "Amount_norm"].sum()
    ic_pct_rev = safe_div(abs(ic_sum), revenue) if revenue != 0 else np.nan

    # Expense ratio (opex + cogs) / revenue
    exp_ratio = safe_div(abs(cogs) + abs(opex), revenue) if revenue != 0 else np.nan

    # Margins
    gross_margin = safe_div(gross_profit, revenue) if revenue != 0 else np.nan
    ebitda_margin = safe_div(ebitda, revenue) if revenue != 0 else np.nan
    net_margin = safe_div(net_profit, revenue) if revenue != 0 else np.nan

    out = {
        "days_in_month": days,
        "Revenue": revenue,
        "COGS": cogs,
        "Opex": opex,
        "Interest": interest,
        "Tax": tax,
        "Gross Profit": gross_profit,
        "EBITDA": ebitda,
        "Net Profit": net_profit,
        "Operating Cash Flow": ocf,
        "Intercompany Sum": ic_sum,
        "Intercompany % of Revenue": ic_pct_rev,
        "Expense Ratio": exp_ratio,
        "Gross Margin %": gross_margin,
        "EBITDA Margin %": ebitda_margin,
        "Net Margin %": net_margin,
        "df_month": df,
    }
    return out


def compute_wc_days(ar, ap, inv, revenue, cogs, days_in_month):
    dso = safe_div(ar, revenue) * days_in_month if revenue else np.nan
    dpo = safe_div(ap, abs(cogs)) * days_in_month if cogs else np.nan
    dio = safe_div(inv, abs(cogs)) * days_in_month if cogs else np.nan
    ccc = (dso if not np.isnan(dso) else 0) + (dio if not np.isnan(dio) else 0) - (dpo if not np.isnan(dpo) else 0)
    # If all missing, return nan
    if np.isnan(dso) and np.isnan(dpo) and np.isnan(dio):
        ccc = np.nan
    return dso, dpo, dio, ccc


def monthly_pnl_table(tx, entity):
    # Build Jan-Dec per available months (not fixed to 2026, but you can filter)
    df = tx[tx["Entity"] == entity].copy()
    months = sorted([m for m in df["Month"].dropna().unique()])
    if not months:
        return pd.DataFrame()

    rows = []
    for m in months:
        k = compute_kpi_for_entity_month(tx, entity, pd.Timestamp(m))
        rows.append(
            {
                "Month": pd.Timestamp(m),
                "Revenue": k["Revenue"],
                "COGS": k["COGS"],
                "Gross Profit": k["Gross Profit"],
                "Opex": k["Opex"],
                "EBITDA": k["EBITDA"],
                "Interest": k["Interest"],
                "Tax": k["Tax"],
                "Net Profit": k["Net Profit"],
            }
        )
    out = pd.DataFrame(rows).sort_values("Month")
    out["Month"] = out["Month"].dt.strftime("%b-%Y")
    return out


def sister_concerns_table(tx, month_ts, revenue_by_entity=None):
    ms, me = month_start_end(month_ts)
    dfm = tx[(tx["Date"] >= ms) & (tx["Date"] <= me)].copy()
    ic = dfm[dfm["IntercompanyFlag"].str.upper().eq("YES")].copy()
    if ic.empty:
        return pd.DataFrame(columns=["Entity", "Inter-company In/Out", "% of Revenue"])

    g = ic.groupby("Entity")["Amount_norm"].sum().reset_index()
    g = g.rename(columns={"Amount_norm": "Inter-company In/Out"})
    if revenue_by_entity is None:
        # compute revenue in same month
        rev = dfm[dfm["AccountType"].str.upper().eq("REVENUE")].groupby("Entity")["Amount_norm"].sum().reset_index()
        rev = rev.rename(columns={"Amount_norm": "Revenue"})
        g = g.merge(rev, on="Entity", how="left")
    else:
        g = g.merge(revenue_by_entity, on="Entity", how="left")

    g["% of Revenue"] = g.apply(lambda r: safe_div(abs(r["Inter-company In/Out"]), r.get("Revenue", np.nan)), axis=1)
    return g


def consolidate_month(tx, month_ts, eliminate_ic=True):
    ms, me = month_start_end(month_ts)
    df = tx[(tx["Date"] >= ms) & (tx["Date"] <= me)].copy()
    if eliminate_ic:
        df = df[~df["IntercompanyFlag"].str.upper().eq("YES")].copy()

    def sum_type(t):
        return df.loc[df["AccountType"].str.upper().eq(t.upper()), "Amount_norm"].sum()

    revenue = sum_type("Revenue")
    cogs = sum_type("COGS")
    opex = sum_type("Opex")
    interest = sum_type("Interest")
    tax = sum_type("Tax")
    gp = revenue + cogs
    ebitda = gp + opex
    npf = ebitda + interest + tax

    exp_ratio = safe_div(abs(cogs) + abs(opex), revenue) if revenue != 0 else np.nan

    return {
        "Revenue": revenue,
        "COGS": cogs,
        "Gross Profit": gp,
        "Opex": opex,
        "EBITDA": ebitda,
        "Interest": interest,
        "Tax": tax,
        "Net Profit": npf,
        "Gross Margin %": safe_div(gp, revenue) if revenue != 0 else np.nan,
        "EBITDA Margin %": safe_div(ebitda, revenue) if revenue != 0 else np.nan,
        "Net Margin %": safe_div(npf, revenue) if revenue != 0 else np.nan,
        "Expense Ratio": exp_ratio,
        "df_month": df,
    }


# -----------------------------
# TEMPLATE CREATOR (EXCEL)
# -----------------------------
def create_blank_kpi_template_xlsx():
    # Creates the EXACT sheets user described, with clean tables.
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    wb.remove(wb.active)

    blue = PatternFill("solid", fgColor="D9EAF7")
    bold = Font(bold=True)
    center = Alignment(horizontal="left", vertical="center")

    # 1) Inputs
    ws = wb.create_sheet("Inputs")
    ws["A1"] = "Business Financial KPI Dashboard – Inputs & Assumptions"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "Control Panel"
    ws["A3"].font = Font(bold=True, size=12)

    inputs_rows = [
        ("Selected Entity", "Entity A"),
        ("Selected Month (YYYY-MM-01)", "2026-01-01"),
        ("AR Closing", 0),
        ("AP Closing", 0),
        ("Inventory Closing", 0),
        ("Cash Balance (month-end)", 0),
        ("Governance Thresholds", ""),
        ("Max Inter-company % of Revenue", 0.20),
        ("Min Cash Balance alert", 100000),
        ("Max Expense Ratio alert", 0.75),
        ("Min Net Margin alert", 0.10),
        ("Max CCC (Days) alert", 60),
        ("SDE/SOE Thresholds", ""),
        ("SDE/SOE Max Expense Ratio", 0.75),
        ("SDE/SOE Min Surplus Margin", 0.10),
        ("SDE/SOE Max Transfer % of Income", 0.35),
    ]
    r = 5
    for k, v in inputs_rows:
        ws[f"A{r}"] = k
        ws[f"B{r}"] = v
        ws[f"B{r}"].fill = blue
        ws[f"A{r}"].font = bold if ("Threshold" in k or "Governance" in k or "SDE/SOE" in k) else Font(bold=False)
        ws[f"A{r}"].alignment = center
        ws[f"B{r}"].alignment = center
        r += 1
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 26

    # 2) Transactions
    ws = wb.create_sheet("Transactions")
    headers = ["Date", "Entity", "AccountType", "Category", "Counterparty", "Amount", "CashFlag", "IntercompanyFlag"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = bold
        cell.fill = PatternFill("solid", fgColor="EEF2F6")
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 16

    # 3) P&L
    wb.create_sheet("P&L")
    # 4) CashFlow
    wb.create_sheet("CashFlow")
    # 5) SisterConcerns
    wb.create_sheet("SisterConcerns")
    # 6) KPI_Dashboard
    wb.create_sheet("KPI_Dashboard")
    # Helper
    wb.create_sheet("P&L_Entity")

    return wb


# -----------------------------
# SIDEBAR – FILE INPUTS + THRESHOLDS
# -----------------------------
st.title("Integrated Financial Performance & Governance Dashboard")

with st.sidebar:
    st.header("Data Sources")

    kpi_upload = st.file_uploader("Upload KPI Template (xlsx)", type=["xlsx"], key="kpi_upl")
    sde_upload = st.file_uploader("Upload SDE/SOE file (xlsx)", type=["xlsx"], key="sde_upl")

    st.divider()
    st.header("Controls")

# Load files (uploaded overrides default)
try:
    kpi_src = kpi_upload if kpi_upload is not None else DEFAULT_KPI_PATH
    tx = load_kpi_template(kpi_src)
    kpi_loaded = True
except Exception as e:
    tx = None
    kpi_loaded = False
    st.error(f"KPI template load failed: {e}")

try:
    sde_src = sde_upload if sde_upload is not None else DEFAULT_SDE_PATH
    sde_tables = load_sde_soe_tables(sde_src)
    sde_loaded = True
except Exception as e:
    sde_tables = None
    sde_loaded = False
    st.error(f"SDE/SOE file load failed: {e}")

# -----------------------------------
# TOP NAV (MORE TABS as requested)
# -----------------------------------
top_tabs = st.tabs(
    [
        "🏠 Home / Model Map",
        "📊 KPI Template Dashboard",
        "🏛 Governance (RAG + Improvements)",
        "🧾 Consolidation (Group View)",
        "📈 KPI Charts + Narration",
        "🎓 SDE/SOE Dashboard",
        "🧩 SDE/SOE Governance + Notes",
        "🗂 Data Tables (Both Files)",
        "⚙️ Tools (Create Template)",
    ]
)

# -----------------------------------
# HOME / MODEL MAP
# -----------------------------------
with top_tabs[0]:
    st.subheader("Overview Picture / Model Map")

    cols = st.columns([2, 1])
    with cols[0]:
        shown = False
        for p in MODEL_MAP_IMG_PATHS:
            if os.path.exists(p):
                st.image(p, use_container_width=True)
                shown = True
                break
        if not shown:
            st.info("Model map image not found in this environment. (Add your screenshot PNG to the project folder and update MODEL_MAP_IMG_PATHS.)")

    with cols[1]:
        card("KPI Template Loaded", "YES" if kpi_loaded else "NO")
        card("SDE/SOE Loaded", "YES" if sde_loaded else "NO")
        st.markdown("---")
        st.markdown(
            """
            **What this dashboard does**
            - Reads your ledger-style **Transactions** and builds P&L, CashFlow, SisterConcerns, KPIs.
            - Applies strict **RED / AMBER / GREEN** governance checks with clear actions.
            - Reads **SDE/SOE/CDOE** year-wise file and produces trend charts + governance notes.
            - Produces **Consolidation** (all entities) with optional intercompany elimination.
            """
        )

# -----------------------------------
# KPI TEMPLATE DASHBOARD
# -----------------------------------
if kpi_loaded:
    entities = sorted(tx["Entity"].dropna().unique().tolist())
    months = sorted([m for m in tx["Month"].dropna().unique()])
else:
    entities, months = [], []

with st.sidebar:
    if kpi_loaded:
        sel_entity = st.selectbox("Selected Entity", options=entities, index=0)
        # Show months as "Jan-2026" etc.
        month_labels = [pd.Timestamp(m).strftime("%b-%Y") for m in months]
        sel_month_label = st.selectbox("Selected Month", options=month_labels, index=0)
        sel_month = pd.Timestamp(months[month_labels.index(sel_month_label)])
    else:
        sel_entity, sel_month = None, None

    st.divider()
    st.subheader("Working Capital Inputs")
    ar = st.number_input("AR Closing (₹)", min_value=0.0, value=450000.0, step=10000.0, format="%.0f")
    ap = st.number_input("AP Closing (₹)", min_value=0.0, value=380000.0, step=10000.0, format="%.0f")
    inv = st.number_input("Inventory Closing (₹)", min_value=0.0, value=250000.0, step=10000.0, format="%.0f")
    cash_bal = st.number_input("Cash Balance (month-end) (₹)", min_value=0.0, value=0.0, step=10000.0, format="%.0f")

    st.divider()
    st.subheader("Governance Thresholds")
    max_ic_pct = st.number_input("Max Inter-company % of Revenue", min_value=0.0, max_value=1.0, value=0.20, step=0.01, format="%.2f")
    min_cash_alert = st.number_input("Min Cash Balance alert (₹)", min_value=0.0, value=100000.0, step=10000.0, format="%.0f")
    max_exp_ratio = st.number_input("Max Expense Ratio alert", min_value=0.0, max_value=2.0, value=0.75, step=0.01, format="%.2f")
    min_net_margin = st.number_input("Min Net Margin alert", min_value=-1.0, max_value=1.0, value=0.10, step=0.01, format="%.2f")
    max_ccc_days = st.number_input("Max CCC (Days) alert", min_value=0.0, value=60.0, step=5.0, format="%.0f")

# Compute KPI
if kpi_loaded and sel_entity and sel_month is not None:
    k = compute_kpi_for_entity_month(tx, sel_entity, sel_month)
    dso, dpo, dio, ccc = compute_wc_days(ar, ap, inv, k["Revenue"], k["COGS"], k["days_in_month"])
else:
    k = None
    dso = dpo = dio = ccc = np.nan

with top_tabs[1]:
    st.subheader("KPI Template Dashboard (Monthly / Ledger-driven)")

    if not kpi_loaded:
        st.warning("Upload the KPI template file to enable this section.")
    else:
        # KPI cards (NO TRUNCATION, FULL ₹ with commas)
        c1, c2, c3, c4, c5, c6 = st.columns(6)
        with c1:
            card("Total Revenue", inr(k["Revenue"]))
        with c2:
            card("Operating Expenses (Opex)", inr(abs(k["Opex"])), "Shown as absolute (cost).")
        with c3:
            card("Net Profit", inr(k["Net Profit"]))
        with c4:
            card("Operating Cash Flow", inr(k["Operating Cash Flow"]))
        with c5:
            card("Expense Ratio", pct(k["Expense Ratio"]))
        with c6:
            card("Inter-company %", pct(k["Intercompany % of Revenue"]))

        st.caption("Sign convention enforced: Revenue (+); COGS/Opex/Interest/Tax are negative; Profit is simple addition.")

        # Sub-tabs for KPI template (matches your logic)
        sub = st.tabs(["Inputs", "Transactions", "P&L", "CashFlow", "SisterConcerns", "KPI Table"])

        # Inputs
        with sub[0]:
            st.markdown("**Purpose:** Single control panel + assumptions (editable inputs are in sidebar).")
            df_inputs = pd.DataFrame(
                [
                    ["Selected Entity", sel_entity],
                    ["Selected Month", sel_month.strftime("%b-%Y")],
                    ["AR Closing", inr(ar)],
                    ["AP Closing", inr(ap)],
                    ["Inventory Closing", inr(inv)],
                    ["Cash Balance (month-end)", inr(cash_bal)],
                    ["Max Inter-company % of Revenue", pct(max_ic_pct)],
                    ["Min Cash Balance alert", inr(min_cash_alert)],
                    ["Max Expense Ratio alert", pct(max_exp_ratio)],
                    ["Min Net Margin alert", pct(min_net_margin)],
                    ["Max CCC (Days) alert", f"{max_ccc_days:.0f}"],
                ],
                columns=["Input", "Value"],
            )
            st.dataframe(df_inputs, use_container_width=True, hide_index=True)

        # Transactions
        with sub[1]:
            st.markdown("**Purpose:** Clean ledger table that drives everything.")
            dfm = k["df_month"].copy()
            show_cols = ["Date", "Entity", "AccountType", "Category", "Counterparty", "Amount", "Amount_norm", "CashFlag", "IntercompanyFlag"]
            dfm = dfm[show_cols].sort_values("Date")
            dfm = dfm.rename(columns={"Amount_norm": "Amount (Sign Enforced)"})
            st.dataframe(dfm, use_container_width=True, hide_index=True)

        # P&L
        with sub[2]:
            st.markdown("**Purpose:** Monthly P&L auto-built from Transactions (SUMIFS-equivalent).")
            pnl = monthly_pnl_table(tx, sel_entity)
            if pnl.empty:
                st.info("No data for this entity.")
            else:
                st.dataframe(pnl, use_container_width=True, hide_index=True)

        # CashFlow
        with sub[3]:
            st.markdown("**Purpose:** Liquidity view + cash conversion cycle.")
            c1, c2, c3, c4 = st.columns(4)
            with c1:
                card("DSO (AR Days)", "—" if np.isnan(dso) else f"{dso:.1f}")
            with c2:
                card("DPO (AP Days)", "—" if np.isnan(dpo) else f"{dpo:.1f}")
            with c3:
                card("DIO (Inv Days)", "—" if np.isnan(dio) else f"{dio:.1f}")
            with c4:
                card("CCC (Days)", "—" if np.isnan(ccc) else f"{ccc:.1f}")

            st.markdown("**Operating Cash Flow (CashFlag = Cash)**")
            st.write(inr(k["Operating Cash Flow"]))

        # SisterConcerns
        with sub[4]:
            st.markdown("**Purpose:** Inter-company monitoring and flags.")
            sc = sister_concerns_table(tx, sel_month)
            if sc.empty:
                st.info("No inter-company rows in selected month.")
            else:
                sc2 = sc.copy()
                sc2["Inter-company In/Out"] = sc2["Inter-company In/Out"].apply(inr)
                sc2["% of Revenue"] = sc2["% of Revenue"].apply(lambda x: pct(x, 1))
                st.dataframe(sc2[["Entity", "Inter-company In/Out", "% of Revenue"]], use_container_width=True, hide_index=True)

        # KPI Table
        with sub[5]:
            st.markdown("**Purpose:** KPI table with margins + working capital days.")
            kpi_tbl = pd.DataFrame(
                [
                    ["Gross Margin %", pct(k["Gross Margin %"])],
                    ["EBITDA Margin %", pct(k["EBITDA Margin %"])],
                    ["Net Profit Margin %", pct(k["Net Margin %"])],
                    ["DSO (AR Days)", "—" if np.isnan(dso) else f"{dso:.1f}"],
                    ["DPO (AP Days)", "—" if np.isnan(dpo) else f"{dpo:.1f}"],
                    ["DIO (Inv Days)", "—" if np.isnan(dio) else f"{dio:.1f}"],
                    ["CCC (Days)", "—" if np.isnan(ccc) else f"{ccc:.1f}"],
                ],
                columns=["KPI", "Value"],
            )
            st.dataframe(kpi_tbl, use_container_width=True, hide_index=True)

# -----------------------------------
# GOVERNANCE (RAG + IMPROVEMENTS)
# -----------------------------------
with top_tabs[2]:
    st.subheader("Governance Suggestions (RED / AMBER / GREEN)")

    if not kpi_loaded:
        st.warning("Upload KPI template to enable governance checks.")
    else:
        # Build checks
        exp_ratio = k["Expense Ratio"]
        net_margin = k["Net Margin %"]
        ic_pct = k["Intercompany % of Revenue"]
        cash_ok = cash_bal
        ccc_days = ccc

        rag_exp, why_exp = rag_from_threshold(exp_ratio, green_max=max_exp_ratio, higher_is_bad=True)
        rag_ic, why_ic = rag_from_threshold(ic_pct, green_max=max_ic_pct, higher_is_bad=True)
        rag_cash, why_cash = rag_from_threshold(cash_ok, green_min=min_cash_alert, higher_is_bad=False)
        rag_nm, why_nm = rag_from_threshold(net_margin, green_min=min_net_margin, higher_is_bad=False)
        rag_ccc, why_ccc = rag_from_threshold(ccc_days, green_max=max_ccc_days, higher_is_bad=True)

        # Action recommendations (strict + practical)
        actions = {
            "Expense Ratio": {
                "RED": "RED: Expense Ratio HIGH. Improve: reduce admin/overheads, renegotiate vendors, stop discretionary spend, re-check COGS leakage, category-wise caps.",
                "AMBER": "AMBER: Expense Ratio near limit. Improve: lock budgets by category, approve PO limits, monthly variance review (Budget vs Actual).",
                "GREEN": "GREEN: Expense Ratio under control.",
            },
            "Intercompany %": {
                "RED": "RED: Inter-company % is HIGH. Improve: enforce IC caps, reconcile IC balances weekly, require approvals for SisterCo payments, align transfer pricing / MoU.",
                "AMBER": "AMBER: Inter-company % near limit. Improve: monitor weekly, require justification & tagging for IC entries.",
                "GREEN": "GREEN: Inter-company level acceptable (as per threshold).",
            },
            "Cash Balance": {
                "RED": "RED: Cash Balance below minimum. Improve: accelerate collections (DSO), defer non-critical payments, renegotiate payment terms, short-term liquidity plan.",
                "AMBER": "AMBER: Cash Balance barely above minimum. Improve: tighten cash forecasting, daily cash dashboard for top 10 inflows/outflows.",
                "GREEN": "GREEN: Cash balance acceptable.",
            },
            "Net Margin": {
                "RED": "RED: Net Margin below minimum. Improve: reprice services, reduce cost base, increase high-margin revenue mix, cut non-core Opex.",
                "AMBER": "AMBER: Net Margin borderline. Improve: focus on fee realization, eliminate low-margin categories, improve utilization.",
                "GREEN": "GREEN: Net margin healthy.",
            },
            "CCC (Days)": {
                "RED": "RED: CCC too high. Improve: reduce AR days (collections), reduce inventory days (stock optimization), extend AP days (supplier terms).",
                "AMBER": "AMBER: CCC near limit. Improve: weekly follow-up on receivables + vendor term review.",
                "GREEN": "GREEN: CCC under control.",
            },
        }

        rag_box(rag_exp, actions["Expense Ratio"][rag_exp] + f" (Current: {pct(exp_ratio)} | Limit: {pct(max_exp_ratio)})")
        rag_box(rag_nm, actions["Net Margin"][rag_nm] + f" (Current: {pct(net_margin)} | Min: {pct(min_net_margin)})")
        rag_box(rag_cash, actions["Cash Balance"][rag_cash] + f" (Current: {inr(cash_bal)} | Min: {inr(min_cash_alert)})")
        rag_box(rag_ic, actions["Intercompany %"][rag_ic] + f" (Current: {pct(ic_pct)} | Max: {pct(max_ic_pct)})")
        rag_box(rag_ccc, actions["CCC (Days)"][rag_ccc] + f" (Current: {'—' if np.isnan(ccc_days) else f'{ccc_days:.1f}'} | Max: {max_ccc_days:.0f})")

        st.markdown("### Improvement Recommendations (Auto)")
        recs = []

        # Category-wise top costs (selected month)
        dfm = k["df_month"].copy()
        top_opex = (
            dfm[dfm["AccountType"].str.upper().eq("OPEX")]
            .groupby("Category")["Amount_norm"]
            .sum()
            .abs()
            .sort_values(ascending=False)
            .head(5)
        )
        if len(top_opex) > 0:
            recs.append(f"Top Opex drivers (month): {', '.join([f'{cat} ({inr(val)})' for cat, val in top_opex.items()])}. Consider caps + approvals for top 2 categories.")

        # Intercompany counterparties
        ic = dfm[dfm["IntercompanyFlag"].str.upper().eq("YES")].copy()
        top_ic = ic.groupby("Counterparty")["Amount_norm"].sum().abs().sort_values(ascending=False).head(5)
        if len(top_ic) > 0:
            recs.append(f"Top intercompany counterparties: {', '.join([f'{cp} ({inr(val)})' for cp, val in top_ic.items()])}. Add IC justification notes + monthly reconciliations.")

        # Working capital flags
        if not np.isnan(dso) and dso > 60:
            recs.append("DSO is high. Do a receivables ageing review; set collection targets and escalation workflow.")
        if not np.isnan(dio) and dio > 90:
            recs.append("Inventory days high. Identify slow-moving items; tighten procurement and reorder points.")
        if not np.isnan(dpo) and dpo < 30:
            recs.append("DPO low. Negotiate vendor terms to reduce working capital pressure.")

        if not recs:
            recs = ["No major red flags detected from available data; keep monthly governance review cadence."]

        for r in recs:
            st.markdown(f"- {r}")

# -----------------------------------
# CONSOLIDATION
# -----------------------------------
with top_tabs[3]:
    st.subheader("Consolidation Report (All Entities - Selected Month)")

    if not kpi_loaded:
        st.warning("Upload KPI template to enable consolidation.")
    else:
        eliminate_ic = st.toggle("Eliminate Intercompany in Consolidation (Recommended)", value=True)
        cons = consolidate_month(tx, sel_month, eliminate_ic=eliminate_ic)

        # Table
        cons_tbl = pd.DataFrame(
            [
                ["Revenue", cons["Revenue"]],
                ["COGS", cons["COGS"]],
                ["Gross Profit", cons["Gross Profit"]],
                ["Opex", cons["Opex"]],
                ["EBITDA", cons["EBITDA"]],
                ["Interest", cons["Interest"]],
                ["Tax", cons["Tax"]],
                ["Net Profit", cons["Net Profit"]],
            ],
            columns=["Metric", "Value"],
        )
        cons_tbl["Value"] = cons_tbl["Value"].apply(inr)
        st.dataframe(cons_tbl, use_container_width=True, hide_index=True)

        # Consolidated KPI cards
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            card("Consolidated Gross Margin", pct(cons["Gross Margin %"]))
        with c2:
            card("Consolidated EBITDA Margin", pct(cons["EBITDA Margin %"]))
        with c3:
            card("Consolidated Net Margin", pct(cons["Net Margin %"]))
        with c4:
            card("Consolidated Expense Ratio", pct(cons["Expense Ratio"]))

        # Quick narrative
        bullets = []
        bullets.append(f"Consolidation basis: **{'IC eliminated' if eliminate_ic else 'IC included'}** for {sel_month.strftime('%b-%Y')}.")
        if not np.isnan(cons["Expense Ratio"]) and cons["Expense Ratio"] > max_exp_ratio:
            bullets.append("Expense Ratio breaches threshold at group level → prioritize cost controls and re-check COGS classification.")
        if not np.isnan(cons["Net Margin %"]) and cons["Net Margin %"] < min_net_margin:
            bullets.append("Net Margin below threshold → review pricing + fee realization + cost structure.")
        narrative_block("Management Summary (Auto)", bullets)

# -----------------------------------
# KPI CHARTS + NARRATION (IMPROVED)
# - NO "stupid" narration; structured, numeric, actionable
# -----------------------------------
with top_tabs[4]:
    st.subheader("KPI Charts + Narration (Selected Entity)")

    if not kpi_loaded:
        st.warning("Upload KPI template to enable charts.")
    else:
        pnl = monthly_pnl_table(tx, sel_entity)
        if pnl.empty:
            st.info("No data for this entity.")
        else:
            # Bar charts (no confusing M scale; values shown above bars)
            pnl_long = pnl.copy()
            pnl_long["Month"] = pd.Categorical(pnl_long["Month"], categories=pnl["Month"].tolist(), ordered=True)

            # Revenue vs Opex vs Net Profit bar
            chart_df = pnl_long[["Month", "Revenue", "Opex", "Net Profit"]].copy()
            chart_df["Opex"] = chart_df["Opex"].abs()
            chart_melt = chart_df.melt(id_vars=["Month"], var_name="Metric", value_name="Value")

            fig = plot_grouped_bar_with_labels(chart_melt, x="Month", y="Value", group="Metric",
                                               title="Monthly: Revenue vs Opex vs Net Profit (Values on bars)",
                                               yaxis_title="₹")
            st.plotly_chart(fig, use_container_width=True)

            # Narration (structured)
            # Current vs previous month
            last_idx = pnl.shape[0] - 1
            cur = pnl.iloc[last_idx].to_dict()
            prev = pnl.iloc[last_idx - 1].to_dict() if last_idx - 1 >= 0 else None

            def delta(a, b):
                if b is None or b == 0 or (isinstance(b, float) and np.isnan(b)):
                    return np.nan
                return (a - b) / abs(b)

            bullets = []
            bullets.append(f"Current month **{pnl.iloc[last_idx]['Month']}**: Revenue {inr(cur['Revenue'])}, Opex {inr(abs(cur['Opex']))}, Net Profit {inr(cur['Net Profit'])}.")
            if prev:
                bullets.append(
                    f"MoM change: Revenue {pct(delta(cur['Revenue'], prev['Revenue']))}, "
                    f"Opex {pct(delta(abs(cur['Opex']), abs(prev['Opex'])) )}, "
                    f"Net Profit {pct(delta(cur['Net Profit'], prev['Net Profit']))}."
                )

            # Cost driver narrative from selected month
            dfm = k["df_month"].copy()
            opex_by_cat = (
                dfm[dfm["AccountType"].str.upper().eq("OPEX")]
                .groupby("Category")["Amount_norm"].sum().abs().sort_values(ascending=False)
            )
            if len(opex_by_cat) > 0:
                top2 = opex_by_cat.head(2)
                bullets.append(
                    "Top Opex drivers (selected month): "
                    + ", ".join([f"**{c}** {inr(v)}" for c, v in top2.items()])
                    + ". Action: cap/approve these categories first."
                )

            # Governance headline
            exp_ratio = k["Expense Ratio"]
            ic_pct = k["Intercompany % of Revenue"]
            nm = k["Net Margin %"]
            bullets.append(f"Governance snapshot: Expense Ratio {pct(exp_ratio)} vs limit {pct(max_exp_ratio)}; IC% {pct(ic_pct)} vs limit {pct(max_ic_pct)}; Net Margin {pct(nm)} vs min {pct(min_net_margin)}.")

            narrative_block("Narration (Clean, numeric, actionable)", bullets)

# -----------------------------------
# SDE/SOE DASHBOARD
# -----------------------------------
with top_tabs[5]:
    st.subheader("SDE/SOE Dashboard")

    if not sde_loaded:
        st.warning("Upload SDE/SOE file to enable this section.")
    else:
        cd = sde_tables["CDOE"].copy()
        comb = sde_tables["COMBINED"].copy()

        # Prefer CDOE sheet for Income/Expenditure/Surplus/Transfer (it has all columns clearly)
        # Standardize column names
        rename_map = {}
        for c in cd.columns:
            cl = c.lower()
            if "enrollment" in cl:
                rename_map[c] = "Enrollment"
            elif cl == "income":
                rename_map[c] = "Income"
            elif "expenditure" in cl:
                rename_map[c] = "Expenditure"
            elif "surplus" in cl:
                rename_map[c] = "Surplus"
            elif "transferred" in cl or "transfer" in cl:
                rename_map[c] = "Transfer_to_BVDU"
        cd = cd.rename(columns=rename_map)

        # Keep only needed
        keep = ["Year", "Enrollment", "Income", "Expenditure", "Surplus", "Transfer_to_BVDU"]
        for c in keep:
            if c not in cd.columns:
                cd[c] = np.nan
        cd2 = cd[keep].copy()
        cd2 = cd2.dropna(subset=["Year"])

        st.markdown("### Year-wise Financial Status (CDOE)")
        st.dataframe(cd2, use_container_width=True, hide_index=True)

        # BAR charts with values above bars (NO confusing M axis)
        # Income vs Expenditure
        df_ie = cd2.melt(id_vars=["Year"], value_vars=["Income", "Expenditure"], var_name="Metric", value_name="Value")
        fig_ie = plot_grouped_bar_with_labels(df_ie, x="Year", y="Value", group="Metric",
                                              title="Income vs Expenditure (Year-wise)",
                                              yaxis_title="₹")
        st.plotly_chart(fig_ie, use_container_width=True)

        # Surplus + Transfer
        df_st = cd2.melt(id_vars=["Year"], value_vars=["Surplus", "Transfer_to_BVDU"], var_name="Metric", value_name="Value")
        fig_st = plot_grouped_bar_with_labels(df_st, x="Year", y="Value", group="Metric",
                                              title="Surplus & Transfer to BVDU (Year-wise)",
                                              yaxis_title="₹")
        st.plotly_chart(fig_st, use_container_width=True)

        # Enrollment
        fig_en = plot_bar_with_labels(cd2, x="Year", y="Enrollment", title="Enrollment (Year-wise)", yaxis_title="Students")
        st.plotly_chart(fig_en, use_container_width=True)

        # Narration (organized)
        # compute ratios latest year
        latest = cd2.dropna(subset=["Income"]).tail(1)
        if not latest.empty:
            r = latest.iloc[0]
            exp_ratio = safe_div(r["Expenditure"], r["Income"])
            surplus_margin = safe_div(r["Surplus"], r["Income"])
            transfer_pct = safe_div(r["Transfer_to_BVDU"], r["Income"])

            bullets = [
                f"Latest year **{r['Year']}**: Income {inr(r['Income'])}, Expenditure {inr(r['Expenditure'])}, Surplus {inr(r['Surplus'])}.",
                f"Expense Ratio = {pct(exp_ratio)}; Surplus Margin = {pct(surplus_margin)}; Transfer % of Income = {pct(transfer_pct)}.",
                "Use the governance tab to see strict RED/AMBER/GREEN flags with action steps.",
            ]
            narrative_block("Narration (SDE/SOE/CDOE)", bullets)

# -----------------------------------
# SDE/SOE GOVERNANCE + NOTES
# -----------------------------------
with top_tabs[6]:
    st.subheader("SDE/SOE Governance Suggestions + Notes / Improvements")

    if not sde_loaded:
        st.warning("Upload SDE/SOE file to enable this section.")
    else:
        # Thresholds (reuse KPI sidebar values if you want separate; here we keep separate defaults)
        with st.sidebar:
            st.divider()
            st.subheader("SDE/SOE Thresholds")
            sde_max_exp_ratio = st.number_input("SDE/SOE Max Expense Ratio", min_value=0.0, max_value=2.0, value=0.75, step=0.01, format="%.2f")
            sde_min_surplus_margin = st.number_input("SDE/SOE Min Surplus Margin", min_value=-1.0, max_value=1.0, value=0.10, step=0.01, format="%.2f")
            sde_max_transfer_pct = st.number_input("SDE/SOE Max Transfer % of Income", min_value=0.0, max_value=1.0, value=0.35, step=0.01, format="%.2f")

        cd = sde_tables["CDOE"].copy()
        # same standardization as earlier
        rename_map = {}
        for c in cd.columns:
            cl = c.lower()
            if "enrollment" in cl:
                rename_map[c] = "Enrollment"
            elif cl == "income":
                rename_map[c] = "Income"
            elif "expenditure" in cl:
                rename_map[c] = "Expenditure"
            elif "surplus" in cl:
                rename_map[c] = "Surplus"
            elif "transferred" in cl or "transfer" in cl:
                rename_map[c] = "Transfer_to_BVDU"
        cd = cd.rename(columns=rename_map)
        keep = ["Year", "Enrollment", "Income", "Expenditure", "Surplus", "Transfer_to_BVDU"]
        for c in keep:
            if c not in cd.columns:
                cd[c] = np.nan
        cd2 = cd[keep].dropna(subset=["Year"]).copy()

        latest = cd2.dropna(subset=["Income"]).tail(1)
        if latest.empty:
            st.info("No usable Income/Expenditure data found for governance checks.")
        else:
            r = latest.iloc[0]
            exp_ratio = safe_div(r["Expenditure"], r["Income"])
            surplus_margin = safe_div(r["Surplus"], r["Income"])
            transfer_pct = safe_div(r["Transfer_to_BVDU"], r["Income"])

            rag_exp, _ = rag_from_threshold(exp_ratio, green_max=sde_max_exp_ratio, higher_is_bad=True)
            rag_sm, _ = rag_from_threshold(surplus_margin, green_min=sde_min_surplus_margin, higher_is_bad=False)
            rag_tr, _ = rag_from_threshold(transfer_pct, green_max=sde_max_transfer_pct, higher_is_bad=True)

            rag_box(rag_exp, f"{rag_exp}: Expense Ratio {pct(exp_ratio)} (Limit {pct(sde_max_exp_ratio)}). Improve: tighten overheads, vendor renegotiation, discretionary controls.")
            rag_box(rag_sm, f"{rag_sm}: Surplus Margin {pct(surplus_margin)} (Min {pct(sde_min_surplus_margin)}). Improve: fee realization, optimize cost per student, reduce leakage.")
            rag_box(rag_tr, f"{rag_tr}: Transfer to BVDU {pct(transfer_pct)} of Income (Max {pct(sde_max_transfer_pct)}). Improve: justify transfer basis, link to performance, document governance approval.")

            st.markdown("### Notes / Improvements (Auto)")
            notes = []
            if rag_exp == "RED":
                notes.append("Immediate: prepare category-wise expense reduction plan; identify top 3 cost heads and set monthly caps.")
            if rag_sm in ("RED", "AMBER"):
                notes.append("Improve surplus: revise pricing / fee realization, raise enrollment quality/volume, reduce cost per student.")
            if rag_tr in ("RED", "AMBER"):
                notes.append("Strengthen governance: add transfer approval matrix + monthly transfer reconciliation note.")
            if not notes:
                notes.append("Governance looks stable for latest year, continue quarterly review and documentation.")
            for n in notes:
                st.markdown(f"- {n}")

# -----------------------------------
# DATA TABLES (BOTH FILES)
# -----------------------------------
with top_tabs[7]:
    st.subheader("Data Tables (Both Files)")

    c1, c2 = st.columns(2)
    with c1:
        st.markdown("### KPI Template – Transactions (Sample)")
        if not kpi_loaded:
            st.info("KPI template not loaded.")
        else:
            st.dataframe(tx.head(200), use_container_width=True, hide_index=True)

    with c2:
        st.markdown("### SDE/SOE – CDOE Table (Sample)")
        if not sde_loaded:
            st.info("SDE/SOE file not loaded.")
        else:
            st.dataframe(sde_tables["CDOE"].head(200), use_container_width=True, hide_index=True)

# -----------------------------------
# TOOLS (CREATE TEMPLATE)
# -----------------------------------
with top_tabs[8]:
    st.subheader("Tools")
    st.markdown("### 1) Auto-create the starter Excel template (exact sheets + columns)")

    save_dir = st.text_input("Save folder (Windows example: D:\\AJ_Budget Analysis)", value=r"D:\AJ_Budget Analysis")
    file_name = st.text_input("Template file name", value="KPI_Template_AutoCreated.xlsx")

    if st.button("Create Excel Template Now"):
        wb = create_blank_kpi_template_xlsx()

        # Save into memory for download
        import io
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)

        # Try saving to disk path (best effort)
        saved_path = None
        try:
            if save_dir and os.path.isdir(save_dir):
                saved_path = os.path.join(save_dir, file_name)
                wb.save(saved_path)
        except Exception:
            saved_path = None

        st.success("Template created.")
        if saved_path:
            st.info(f"Saved to: {saved_path}")

        st.download_button(
            "Download Template",
            data=bio.getvalue(),
            file_name=file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    st.markdown("---")
    st.markdown("### 2) Quick sanity checks (why amounts looked wrong earlier)")
    st.markdown(
        """
        **Common causes**
        - Revenue entered as negative (we now force Revenue to positive).
        - Costs entered positive (we now force COGS/Opex/Interest/Tax to negative).
        - Month filter mismatch (we use Date → MonthStart/MonthEnd).
        - Truncated KPI cards (we now format full ₹ values with commas, no '...').
        """
    )